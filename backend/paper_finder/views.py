import uuid, time, csv
from typing import Dict, List, Tuple
from collections import defaultdict


from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseBadRequest
from django.core.paginator import Paginator
from django.views.decorators.http import require_POST, require_GET
from django.urls import reverse
from django.utils.encoding import smart_str
from django.db.models import Count
from django.http import JsonResponse
from django.db import transaction, IntegrityError

from .services.matching import keep_paper_boolean, aff_hits
from .services.storage import dedupe_in_memory, save_items
from .utils.names import generate_variants, all_variant_texts_for_match

from .forms import SearchForm, PaperFilterForm
from .models import SearchJob, Paper
from .sources.openalex import OpenAlexAdapter
from .sources.crossref import CrossrefAdapter
from .sources.arxiv import ArxivAdapter
from .sources.scholar import ScholarAdapter


# Celery 可选
try:
    from .tasks import run_job_async
    HAS_CELERY = True
except Exception:
    HAS_CELERY = False


ADAPTERS = {
    "openalex": OpenAlexAdapter(),
    "crossref": CrossrefAdapter(mailto="user@example.com"),
    "arxiv": ArxivAdapter(),
    "scholar": ScholarAdapter(),
}

# --- 批量解析工具 ---
import io, csv
try:
    from openpyxl import load_workbook  # 如不需要Excel可删
    HAS_XLSX = True
except Exception:
    HAS_XLSX = False

def _parse_names_text(s: str):
    """
    每行：姓名[, 拼音][, 单位]
    """
    rows = []
    for raw in (s or "").splitlines():
        line = raw.strip()
        if not line:
            continue
        parts = [p.strip() for p in line.split(",")]
        rows.append({
            "name": parts[0] if len(parts) >= 1 else "",
            "pinyin": parts[1] if len(parts) >= 2 else "",
            "affiliation": parts[2] if len(parts) >= 3 else "",
        })
    return [r for r in rows if r["name"]]

def _parse_csv(fileobj):
    content = fileobj.read()
    text = None
    for enc in ("utf-8-sig", "utf-8", "gbk"):
        try:
            text = content.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("CSV 编码无法识别，请使用 UTF-8。")
    reader = csv.DictReader(io.StringIO(text))
    out = []
    for r in reader:
        out.append({
            "name": (r.get("name") or "").strip(),
            "pinyin": (r.get("pinyin") or "").strip(),
            "affiliation": (r.get("affiliation") or "").strip(),
            "start_date": (r.get("start_date") or "").strip(),
            "end_date": (r.get("end_date") or "").strip(),
        })
    return [x for x in out if x["name"]]

def _parse_excel(fileobj):
    if not HAS_XLSX:
        raise ValueError("缺少 openpyxl 依赖，无法解析 Excel")
    wb = load_workbook(filename=io.BytesIO(fileobj.read()), data_only=True)
    ws = wb.active
    headers = [(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    def idx(col): return headers.index(col) if col in headers else None
    i_name = idx("name"); i_py = idx("pinyin"); i_aff = idx("affiliation"); i_s = idx("start_date"); i_e = idx("end_date")
    out = []
    for row in ws.iter_rows(min_row=2):
        def get(i):
            if i is None: return ""
            v = row[i].value
            return (str(v).strip()) if v is not None else ""
        item = {"name": get(i_name), "pinyin": get(i_py), "affiliation": get(i_aff),
                "start_date": get(i_s), "end_date": get(i_e)}
        if item["name"]:
            out.append(item)
    return out


# === 内部工具：把姓名变体结果统一成 [(priority:int, text:str)] ===
def _normalize_variants(raw) -> List[Tuple[int, str]]:
    pairs: List[Tuple[int, str]] = []
    if isinstance(raw, dict):
        # 形如 {0:[...],1:[...]}
        for pri, items in raw.items():
            p = int(pri) if str(pri).isdigit() else 1
            for t in items or []:
                t = (t or "").strip()
                if t:
                    pairs.append((p, t))
    elif isinstance(raw, list):
        if raw and hasattr(raw[0], "priority"):
            # 老版对象：有 .priority / .text
            for v in raw:
                p = int(getattr(v, "priority", 1))
                t = getattr(v, "text", str(v)).strip()
                if t:
                    pairs.append((p, t))
        else:
            # 简单字符串列表
            for t in raw:
                t = (t or "").strip()
                if t:
                    pairs.append((1, t))
    return pairs


def normalize_doi(doi: str | None) -> str | None:
    if not doi:
        return None
    d = doi.strip()
    d = d.replace("https://doi.org/", "").replace("http://doi.org/", "")
    return d.lower()

@transaction.atomic
def upsert_paper(item, task_ref=None, hints=None):
    """
    item: 一条论文字典，如 {"title":..., "doi":..., "year":..., "authors":..., "source":..., "source_id":...}
    """
    doi = normalize_doi(item.get("doi"))

    # 统一准备 defaults（不要把用于 lookup 的字段再重复放进 defaults 里造成混淆）
    defaults = {
        "title": item.get("title") or "",
        "year": item.get("year"),
        "authors": item.get("authors", ""),
        "source": item.get("source", ""),
        # 你模型的其他字段也在这里补齐
        # "abstract": item.get("abstract", ""),
        # "venue": item.get("venue", ""),
    }

    try:
        if doi:
            # ① 有 DOI：用 DOI 作为唯一查找键
            obj, created = Paper.objects.update_or_create(
                doi=doi,
                defaults=defaults,
            )
        else:
            # ② 没 DOI：退回到你的“来源唯一键”（示例用 source+source_id）
            #    如果你没有 source_id，请改成你自己的指纹/唯一键
            source = item.get("source") or ""
            source_id = item.get("source_id") or ""
            if not source_id:
                # 还可以选择使用 fingerprint(title+year+第一作者) 作为兜底唯一键
                # 这里给个最简单的兜底，避免写入完全不可判重的记录
                source_id = f"no-doi::{(item.get('title') or '').strip().lower()}::{item.get('year') or ''}"
            obj, created = Paper.objects.update_or_create(
                source=source,
                source_id=source_id,
                defaults=defaults | {"doi": None},  # 没 DOI 就显式存 None
            )

    except IntegrityError:
        # ③ 兜底：大概率是并发/历史脏数据导致的 DOI 冲突
        if doi:
            obj = Paper.objects.get(doi=doi)
            # 用 defaults 更新已有记录
            for k, v in defaults.items():
                setattr(obj, k, v)
            obj.save(update_fields=list(defaults.keys()))
        else:
            raise

    return obj

def search_page(request):
    if request.method == "POST":
        # ① 注意要传入 request.FILES
        form = SearchForm(request.POST, request.FILES)
        action = request.POST.get("action", "search")
        if form.is_valid():
            # ---- 先判断是否批量 ----
            names_text = form.cleaned_data.get("names_text")
            names_file = request.FILES.get("names_file")

            if names_text or names_file:
                # 解析
                rows = []
                if names_text:
                    rows += _parse_names_text(names_text)
                if names_file:
                    fname = (names_file.name or "").lower()
                    names_file.seek(0)
                    if fname.endswith(".csv"):
                        rows += _parse_csv(names_file)
                    elif fname.endswith(".xlsx") or fname.endswith(".xlsm"):
                        rows += _parse_excel(names_file)
                    else:
                        # 默认按 CSV 尝试
                        rows += _parse_csv(names_file)

                # 套用“默认值”（来自表单的 default_* 字段）
                def_aff = form.cleaned_data.get("default_affiliation") or ""
                def_start = form.cleaned_data.get("default_start_date")
                def_end   = form.cleaned_data.get("default_end_date")
                for r in rows:
                    r["affiliation"] = r.get("affiliation") or def_aff
                    r["start_date"]  = r.get("start_date") or def_start
                    r["end_date"]    = r.get("end_date") or def_end

                # 去重（按 name+pinyin+aff）
                seen, uniq = set(), []
                for r in rows:
                    key = (r["name"], r.get("pinyin",""), r.get("affiliation",""))
                    if key in seen:
                        continue
                    seen.add(key)
                    uniq.append(r)

                # 为批量中每个姓名创建一个独立 SearchJob，并触发检索
                sources = form.cleaned_data["sources"]
                run_sync = form.cleaned_data.get("run_sync")

                created_job_ids = []
                for r in uniq:
                    name   = r["name"].strip()
                    pinyin = (r.get("pinyin") or "").strip() or None
                    sd = r.get("start_date")
                    ed = r.get("end_date")
                    aff_kw = (r.get("affiliation") or "").strip()
                    hints = {
                        "sources": sources,
                        "query_name": name,
                        "pinyin": pinyin,
                        "aff_kw": aff_kw or None,
                        "date_range": {
                            "start": sd.isoformat() if hasattr(sd, "isoformat") else (sd or None),
                            "end":   ed.isoformat() if hasattr(ed, "isoformat") else (ed or None),
                        },
                    }
                    job_id = uuid.uuid4().hex[:16]
                    SearchJob.objects.create(job_id=job_id, hints=hints, status="running", progress=0.0)
                    created_job_ids.append(job_id)

                    if run_sync or not HAS_CELERY:
                        # 同步：逐个名字跑
                        _run_search_once(job_id, name, pinyin, hints, max_variants=8, pause=0.6)
                    else:
                        # 异步：投 Celery
                        run_job_async.delay(job_id, name, hints, pinyin_override=pinyin)

                # 记录最近一个
                if created_job_ids:
                    request.session["last_job_id"] = created_job_ids[-1]

                # 跳到“历史任务”或直接到结果页（任选其一）
                return redirect("paper_finder:job_history")
                # 或：return redirect(f"{reverse('paper_finder:paper_list')}?job_id={created_job_ids[-1]}")

            # ---- 否则走“单人检索”原逻辑 ----
            name   = form.cleaned_data["name"].strip()
            pinyin = (form.cleaned_data["pinyin"] or "").strip() or None
            sources = form.cleaned_data["sources"]
            sd = form.cleaned_data.get("start_date")
            ed = form.cleaned_data.get("end_date")
            aff_kw = (form.cleaned_data.get("affiliation") or "").strip()

            hints: Dict = {
                "sources": sources,
                "query_name": name,
                "pinyin": pinyin,
                "aff_kw": aff_kw or None,
                "date_range": {
                    "start": sd.isoformat() if sd else None,
                    "end":   ed.isoformat() if ed else None,
                },
            }

            if action == "preview":
                raw = generate_variants(name, pinyin_override=pinyin)
                pairs = _normalize_variants(raw)
                groups = defaultdict(list)
                for pri, text in pairs:
                    groups[int(pri)].append(text)
                variant_count = len(pairs)
                return render(request, "paper_finder/search.html", {
                    "form": form,
                    "variant_count": variant_count,
                    "variant_groups": [
                        (0, groups.get(0, [])),
                        (1, groups.get(1, [])),
                        (2, groups.get(2, [])),
                        (3, groups.get(3, [])),
                    ],
                })

            run_sync = form.cleaned_data.get("run_sync")
            job_id = uuid.uuid4().hex[:16]
            SearchJob.objects.create(job_id=job_id, hints=hints, status="running", progress=0.0)
            request.session["last_job_id"] = job_id

            if run_sync or not HAS_CELERY:
                _run_search_once(job_id, name, pinyin, hints, max_variants=8, pause=0.6)
                return redirect("paper_finder:job_status", job_id=job_id)
            else:
                run_job_async.delay(job_id, name, hints, pinyin_override=pinyin)
                return redirect("paper_finder:job_status", job_id=job_id)

        # 表单校验失败
        return render(request, "paper_finder/search.html", {"form": form})
    else:
        form = SearchForm()
        return render(request, "paper_finder/search.html", {"form": form})


def job_status(request, job_id: str):
    job = get_object_or_404(SearchJob, pk=job_id)
    papers = Paper.objects.filter(task_ref=job_id).order_by("-score","-year")[:50]
    auto_refresh = (job.status == "running")
    return render(request, "paper_finder/job_status.html", {
        "job": job, "papers": papers, "auto_refresh": auto_refresh
    })

@login_required
def paper_list(request):
    job_id = request.GET.get("job_id")

    # 支持多任务对比：jobs=id1,id2 或 ?jobs=id1&jobs=id2
    jobs_param = request.GET.getlist("jobs") or request.GET.get("jobs")
    job_ids = []
    if isinstance(jobs_param, list):
        job_ids = jobs_param
    elif isinstance(jobs_param, str) and jobs_param.strip():
        job_ids = [x.strip() for x in jobs_param.split(",") if x.strip()]

    # --- 兜底：没有任何参数时的处理 ---
    if not job_id and not job_ids:
        # 1) 优先跳转到最近一次任务（存于 session）
        last = request.session.get("last_job_id")
        if last:
            return redirect(f"{reverse('paper_finder:paper_list')}?job_id={last}")

        # 2) 没有 session，就查数据库里最新的一条任务
        latest = SearchJob.objects.order_by("-created_at").first()
        if latest:
            return redirect(f"{reverse('paper_finder:paper_list')}?job_id={latest.job_id}")

        # 3) 仍然没有任何任务 → 展示选择页（空则引导去发起）
        recent_jobs = SearchJob.objects.order_by("-created_at")[:50]
        if not recent_jobs:
            return render(request, "paper_finder/empty_list.html", {
                "tip": "暂无历史结果，请先发起一次检索。"
            })
        return render(request, "paper_finder/paper_list_landing.html", {
            "recent_jobs": recent_jobs
        })

    # --- 有参数时：正常查询 ---
    qs = Paper.objects.all().order_by("-score", "-year", "-id")
    if job_id:
        qs = qs.filter(task_ref=job_id)
    if job_ids:
        qs = qs.filter(task_ref__in=job_ids)

    form = PaperFilterForm(request.GET or None)
    if form.is_valid():
        q = form.cleaned_data.get("q")
        state = form.cleaned_data.get("state")
        source = form.cleaned_data.get("source")
        y1 = form.cleaned_data.get("year_from")
        y2 = form.cleaned_data.get("year_to")
        if q: qs = qs.filter(title__icontains=q)
        if state: qs = qs.filter(state=state)
        if source: qs = qs.filter(source=source)
        if y1: qs = qs.filter(year__gte=y1)
        if y2: qs = qs.filter(year__lte=y2)

    paginator = Paginator(qs, 20)
    page_obj = paginator.get_page(request.GET.get("page") or 1)

    # 供页头的“任务选择框”使用
    recent_jobs = SearchJob.objects.order_by("-created_at")[:100]

    return render(request, "paper_finder/paper_list.html", {
        "form": form,
        "page_obj": page_obj,
        "job_id": job_id,
        "papers": page_obj.object_list,
        "job_ids": ",".join(job_ids) if job_ids else "",
        "recent_jobs": recent_jobs,
    })


def job_history(request):
    # 基础列表（可按时间/关键词过滤）
    qs = SearchJob.objects.all().order_by("-created_at")

    # 轻量过滤（可选）
    q = request.GET.get("q", "").strip()
    if q:
        # 在 hints 的 query_name / aff_kw 上做包含匹配
        qs = qs.filter(hints__query_name__icontains=q) | qs.filter(hints__aff_kw__icontains=q)

    paginator = Paginator(qs, 20)
    page_obj = paginator.get_page(request.GET.get("page") or 1)

    # 统计每个 job 的结果条数（总数与 confirmed 数）
    from django.db.models import Q  # 局部导入，避免全局命名污染
    agg = (
        Paper.objects
        .values("task_ref")
        .annotate(
            total=Count("id"),
            confirmed=Count("id", filter=Q(state="confirmed")),
        )
    )
    stat_map = {a["task_ref"]: (a["total"], a["confirmed"]) for a in agg}

    # 给分页内的每个 job 附上计数
    for job in page_obj.object_list:
        total, confirmed = stat_map.get(job.job_id, (0, 0))
        job.paper_total = total
        job.paper_confirmed = confirmed

    # 记录最近 job
    if page_obj.object_list:
        request.session["last_job_id"] = page_obj.object_list[0].job_id

    return render(request, "paper_finder/job_history.html", {
        "page_obj": page_obj,
        "q": q,
    })


ALLOWED_STATES = {"pending", "confirmed", "rejected"}


@require_POST
def mark_paper(request, pk: int):
    state = (request.POST.get("state") or "").strip().lower()
    if state not in ALLOWED_STATES:
        return HttpResponseBadRequest("invalid state")

    p = get_object_or_404(Paper, pk=pk)
    p.state = state
    p.save(update_fields=["state"])

    # 优先回到提交前的页面
    next_url = request.POST.get("next")
    if next_url:
        return redirect(next_url)

    # 其次带着 job_id 回到结果列表
    job_id = request.POST.get("job_id") or p.task_ref
    if job_id:
        return redirect(f"{reverse('paper_finder:paper_list')}?job_id={job_id}")

    # 最后兜底
    return redirect("paper_finder:paper_list")


def export_csv(request):
    job_id = request.GET.get("job_id")
    state = request.GET.get("state")
    qs = Paper.objects.all()
    if job_id: qs = qs.filter(task_ref=job_id)
    if state: qs = qs.filter(state=state)

    resp = HttpResponse(content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="papers_{job_id or "all"}.csv"'
    w = csv.writer(resp)
    w.writerow(["Title","Year","Venue","DOI","URL","Source","Score","State","Task"])
    for p in qs.iterator(chunk_size=1000):
        w.writerow([p.title, p.year or "", p.venue, p.doi or "", p.url, p.source, f"{p.score:.2f}", p.state, p.task_ref])
    return resp


# --- 内部：同步执行（调试/无 Celery）
def _run_search_once(job_id: str, name: str, pinyin_override: str | None, hints: dict, max_variants=8, pause=0.6):
    raw_variants = generate_variants(name, pinyin_override=pinyin_override)
    pairs = _normalize_variants(raw_variants)
    # 仅取 P<=2 的变体，并限制数量
    pairs.sort(key=lambda x: x[0])
    variants: List[str] = [t for p, t in pairs if p <= 2][:max_variants]

    job = SearchJob.objects.get(pk=job_id)
    sources = hints.get("sources", ["openalex"])
    total = len(variants) * len(sources)
    done = 0

    for v in variants:
        for src in sources:
            adp = ADAPTERS.get(src)
            if not adp:
                continue
            try:
                items = adp.search(v, hints)
            except Exception:
                items = []

            for it in items:
                # 可按需调用 keep_paper_boolean/aff_hits 做筛选
                upsert_paper(it, task_ref=job_id, hints=hints)

            done += 1
            job.progress = round(done / max(1, total), 3)
            job.save(update_fields=["progress"])
            time.sleep(pause)

    job.status = "done"
    job.save(update_fields=["status"])


@require_GET
def export_job_csv(request):
    job_id = request.GET.get("job_id")
    if not job_id:
        return HttpResponse("缺少 job_id", status=400)

    qs = Paper.objects.filter(task_ref=job_id).order_by("-year", "-month", "title")

    resp = HttpResponse(content_type="text/csv; charset=utf-8-sig")
    resp["Content-Disposition"] = f'attachment; filename="{job_id}.csv"'

    writer = csv.writer(resp)
    writer.writerow(["论文题目", "论文全体作者", "学术会议或刊物名称", "出版年份", "出版月", "来源", "DOI", "URL"])
    for p in qs:
        authors_str = ", ".join(p.authors or []) if isinstance(p.authors, list) else (p.authors or "")
        writer.writerow([
            smart_str(p.title or ""),
            smart_str(authors_str),
            smart_str(p.venue or ""),
            p.year or "",
            p.month or "",
            p.source or "",
            smart_str(p.doi or ""),
            smart_str(p.url or ""),
        ])
    return resp

def job_status(request, job_id: str):
    job = get_object_or_404(SearchJob, pk=job_id)
    papers = Paper.objects.filter(task_ref=job_id).order_by("-score", "-year")[:50]
    auto_refresh = (job.status == "running")
    job_percent = int(round(100 * float(job.progress or 0)))  # ✅ 计算百分比，避免模板里做运算
    return render(request, "paper_finder/job_status.html", {
        "job": job,
        "papers": papers,
        "auto_refresh": auto_refresh,
        "job_percent": job_percent,  # ✅ 传给模板
    })

@require_GET
def job_progress(request, job_id: str):
    """轮询接口：返回当前任务进度（0~1）与状态"""
    job = get_object_or_404(SearchJob, pk=job_id)
    pct = int(round(100 * float(job.progress or 0)))
    return JsonResponse({
        "job_id": job.job_id,
        "status": job.status,            # 例如 running / done / failed
        "progress": float(job.progress or 0),
        "percent": pct,                  # 0~100 的整数
    })

from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import login as auth_login
from django.shortcuts import render, redirect

def signup(request):
    if request.method == "POST":
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            auth_login(request, user)
            return redirect("paper_finder:search")   # 登录后去检索页
    else:
        form = UserCreationForm()
    return render(request, "registration/signup.html", {"form": form})
